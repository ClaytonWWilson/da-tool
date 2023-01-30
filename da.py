# import argparse
from colorama import init, Fore, Back, Style
import io
from openpyxl import Workbook
import os
from pick import pick
from PIL import Image
import re
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time

# FEATURE: Ask to keep or delete any existing data
# FEATURE: Run without interactions using argparse

# Constants and globals
KNET_SEARCH_PAGE = "https://knet.csod.com/admin/Users.aspx?Reload=TRUE"
KNET_WELCOME_PAGE = "https://knet.csod.com/LMS/catalog/Welcome.aspx"
TRANSPORTER_ID_REGEX = re.compile("^[0-9]*[A-Z]+[0-9]+[A-Z0-9]+$")
EMPLOYEE_ID_REGEX = re.compile("^[0-9]+$")
USERNAME_REGEX = re.compile("^[a-z]+$")
DSP_MAP = {
    # Enter DSP Name-Code pairs into here
    # Ex. DSP Company One with code DSPO should be
    # "DSP Company One": "DSPO",
}

knet_logged_in = False


def get_input_file():
    SCRIPT_NAME = __file__.split("\\")[-1]
    cwd = os.getcwd()
    folder_contents = os.listdir(cwd)
    file_blacklist = [
        SCRIPT_NAME,
        "run.bat",
        "install.bat",
        "How to install and use.pdf"
    ]
    files = []

    for element in folder_contents:
        if os.path.isfile(os.path.join(cwd, element)) and not element.startswith(".") and not element.startswith("~") and not element in file_blacklist:
            files.append(element)

    if len(files) == 0:
        print_message("Looks like you haven't moved any files with DA info into this folder yet. Please do that before running the script. You can use either a list of employee IDs, usernames, and transporter IDs, or a training roster file from the scheduling tool.", error=True)
        quit()
    TITLE = "Which file contains the DA information? (Use ENTER to continue):\nIf you don't see your file listed, make sure you move it into the same folder as this script, then try running again."

    selected = pick(files, TITLE)

    input_file = os.path.join(cwd, selected[0])

    return input_file


def get_output_file():
    output_file = input(
        "Enter the filename that you would like data to be saved in: ")

    if not output_file.endswith(".xlsx"):
        output_file += ".xlsx"

    output_file = os.path.join(os.getcwd(), output_file)
    return output_file


def get_args_from_menu():
    ARGS = {
        "input_file": "",
        "output_file": "",
        "knet_link": False,
        "da_name": False,
        "transporter_id": False,
        "amc_link": False,
        "employee_id": False,
        "username": False,
        "email": False,
        "dsps": False,
        "onboarding_status": False,
        "reset_knet_pass": False,
        "new_knet_pass": "",
        "photos": False,
    }

    saving_data = False

    # parser = argparse.ArgumentParser()
    # parser.add_argument(
    #     "filename", help="Name of the xlsx file that contains user IDs and transporter IDs")
    # args = parser.parse_args()

    # if os.path.exists(args.filename):
    #     ARGS["input_file"] = os.path.abspath(args.filename)
    # elif os.path.exists(os.path.join(os.getcwd(), args.filename)):
    #     ARGS["input_file"] = os.path.join(os.getcwd(), args.filename)
    # else:
    #     message = "Could not find the input file at\n\"{}\" or\n\"{}\"\nMake sure you typed the location correctly.\n".format(
    #         args.filename, os.path.join(os.getcwd(), args.filename))
    #     print_message(message, error=True)
    #     quit()

    ARGS["input_file"] = get_input_file()

    TITLE = "What would you like to do? (Use SPACE to select and ENTER to continue):"
    OPTIONS = [
        "Everything",
        "Save Knet Link",
        "Save Names",
        "Save Transporter IDs",
        "Save AMConsole Link",
        "Save Employee IDs",
        "Save Usernames",
        "Save Emails",
        "Save DSPs",
        "Save Onboarding Status",
        "Reset Knet Passwords",
        "Download Badge Photos",
    ]

    selected = pick(OPTIONS, TITLE, multiselect=True, min_selection_count=1)

    if ("Everything", 0) in selected:
        for key in ARGS:
            if ARGS[key] == False:
                ARGS[key] = True
        ARGS["new_knet_pass"] = input("New Knet Password: ").strip()
        saving_data = True
    else:
        if ("Save Knet Link", 1) in selected:
            ARGS["knet_link"] = True
            saving_data = True
        if ("Save Names", 2) in selected:
            ARGS["da_name"] = True
            saving_data = True
        if ("Save Transporter IDs", 3) in selected:
            ARGS["transporter_id"] = True
            saving_data = True
        if ("Save AMConsole Link", 4) in selected:
            ARGS["amc_link"] = True
            saving_data = True
        if ("Save Employee IDs", 5) in selected:
            ARGS["employee_id"] = True
            saving_data = True
        if ("Save Usernames", 6) in selected:
            ARGS["username"] = True
            saving_data = True
        if ("Save Emails", 7) in selected:
            ARGS["email"] = True
            saving_data = True
        if ("Save DSPs", 8) in selected:
            ARGS["dsps"] = True
            saving_data = True
        if ("Save Onboarding Status", 9) in selected:
            ARGS["onboarding_status"] = True
            saving_data = True
        if ("Reset Knet Passwords", 10) in selected:
            ARGS["reset_knet_pass"] = True
            ARGS["new_knet_pass"] = input("New Knet Password: ").strip()
        if ("Download Badge Photos", 11) in selected:
            ARGS["photos"] = True

    # Only ask for output file if data needs to be saved
    if saving_data:
        ARGS["output_file"] = get_output_file()

    return ARGS


def get_knet_data_and_change_password(wdriver, driver, ARGS):
    data_store = {}
    data_store.update(driver)
    global knet_logged_in

    if not knet_logged_in:
        wdriver.get(KNET_SEARCH_PAGE)

        print_message("Waiting for Knet Login")
        try:
            while KNET_WELCOME_PAGE not in wdriver.current_url:
                time.sleep(3)
        except KeyboardInterrupt:
            wdriver.quit()
            quit()

        print_message("Login Successful")
        knet_logged_in = True

    wdriver.get(KNET_SEARCH_PAGE)

    try:
        # Clear any data left in user id and username fields
        wdriver.find_element("id", "userIdText").clear()
        wdriver.find_element("id", "userNameText").clear()

        if data_store["employee_id"]:
            wdriver.find_element("id", "userIdText").send_keys(
                data_store["employee_id"])  # Type employee id into search box
        elif data_store["username"]:
            wdriver.find_element("id", "userNameText").send_keys(
                data_store["username"])  # Type username into search box
        else:
            print_message(
                "Employee id and username not present in the input file. Skipping Knet data", error=True, driver_data=data_store)
            return data_store

        wdriver.find_element(
            "xpath", "//td/div/a/span/b").click()  # Click Search
    except Exception as e:
        if isinstance(e, KeyboardInterrupt):
            wdriver.quit()
            quit()
        else:
            print_message("There was an error searching for employee",
                          error=True, driver_data=driver)
            return data_store

    # Get their name
    try:
        data_store["da_name"] = " ".join(wdriver.find_element(
            "xpath", "//tbody/tr/td/a").text.split(", ")[::-1])
    except Exception:
        print_message("Driver could not be found in Knet",
                      error=True, driver_data=driver)
        return data_store

    # Get their username
    if not data_store["username"]:
        data_store["username"] = wdriver.find_element(
            "xpath", "//table[@id=\"resultTable\"]/tbody/tr/td[2]").text.lower()

    # Get their id
    if not data_store["employee_id"]:
        data_store["employee_id"] = wdriver.find_element(
            "xpath", "//table[2]/tbody/tr/td[2]/table/tbody/tr/td[3]").text

    # Try to change password
    if ARGS["reset_knet_pass"]:
        wdriver.find_element(
            "xpath", "//span[@id=\"rptUsers_ctl00_ddlUserOptions\"]/a").click()  # Click dropdown
        # Click Change Password
        wdriver.find_element(
            "xpath", "//*[@id=\"rptUsers_ctl00_ddlUserOptions_lnkPasswordChange\"]").click()

        # Keep trying to click temporary password for up to 10 seconds
        # Otherwise skip changing the password.
        before = time.time()
        while time.time() - before < 10:
            try:
                # Click Temporary password
                wdriver.find_element("id", "passwdReset-manual").click()
                wdriver.find_element(
                    "xpath", "//div[@id=\"dlgPasswdReset\"]/div/div/a/span/b").click()  # Click OK
                break

            except Exception as e:
                if type(e) == KeyboardInterrupt:
                    wdriver.quit()
                    quit()

                print_message(
                    "Can't find the change password button. Retrying...", driver_data=data_store)
                time.sleep(3)

        if time.time() - before >= 10:
            print_message("Password was not changed",
                          error=True, driver_data=data_store)

        CHNG_PWD_URL = wdriver.current_url

        try:
            wdriver.find_element("id", "newPasswordTextBox").send_keys(
                ARGS["new_knet_pass"])  # Enter new password
            wdriver.find_element("id", "confirmPasswordTextBox").send_keys(
                ARGS["new_knet_pass"])  # Enter new password

            wdriver.find_element(
                "xpath", "//a[@id=\"saveImageButton\"]/span/b").click()  # Click Save
        except Exception as e:
            if isinstance(e, KeyboardInterrupt):
                wdriver.quit()
                quit()
            else:
                print_message(
                    "Could not find the password or confirm password text boxes", error=True, driver_data=data_store)

        if wdriver.current_url == CHNG_PWD_URL:  # Check if the new password worked or not
            print_message("Password could not be set to {}".format(
                ARGS["new_knet_pass"]), error=True, driver_data=data_store)
            wdriver.get(KNET_SEARCH_PAGE)
        else:
            print_message("Password set to {}".format(
                ARGS["new_knet_pass"]), driver_data=data_store)

    # Save the link to their Knet profile
    if ARGS["knet_link"]:
        try:
            wdriver.find_element(
                "xpath", "//span[@id=\"rptUsers_ctl00_ddlUserOptions\"]/a").click()  # Click dropdown
            # Click View Transcript
            wdriver.find_element(
                "xpath", "//*[@id=\"rptUsers_ctl00_ddlUserOptions_lnkTranscript\"]").click()
            data_store["knet_link"] = wdriver.current_url
        except Exception as e:
            if isinstance(e, KeyboardInterrupt):
                wdriver.quit()
                quit()
            else:
                print_message(
                    "Couldn't copy knet profile link. Skipping...", error=True, driver_data=data_store)
                data_store["knet_link"] = ""

    return data_store


def get_amc_data(wdriver, driver_data, ARGS):
    before = time.time()
    data_store = {}
    data_store.update(driver_data)

    # Check if transporter id is present
    if not data_store["transporter_id"]:
        print_message(
            "Transporter ID is not present in the input file. Skipping AM Console data...", error=True, driver_data=data_store)
        return data_store

    data_store["amc_link"] = "https://logistics.amazon.com/amconsole/transporter/{}".format(
        data_store["transporter_id"])  # Save link to their AMC profile

    wdriver.get(data_store["amc_link"])

    # Find and format name
    first_name = ""
    last_name = ""
    try:
        first_name = wdriver.find_element(
            "xpath", "//page-profile/div/div[3]/ng-transclude/p[1]").text.replace("First Name: ", "").strip()
        last_name = wdriver.find_element(
            "xpath", "//page-profile/div/div[3]/ng-transclude/p[3]").text.replace("Last Name: ", "").strip()

        data_store["da_name"] = "{} {}".format(first_name, last_name)
    except Exception as e:
        if isinstance(e, KeyboardInterrupt):
            wdriver.quit()
            quit()

        # If it times out here, most likely the driver was not found in AM Console, so we set implicit wait to 0 to save time
        if isinstance(e, NoSuchElementException):
            wdriver.implicitly_wait(0)

        print_message(
            "Can't find first and last name. Skipping...", error=True, driver_data=data_store)

    # Find and format email
    try:
        data_store["email"] = wdriver.find_element(
            "xpath", "//page-profile/div/div[4]/ng-transclude/p[3]").text.replace("Email Address: ", "").strip()
    except Exception as e:
        if isinstance(e, KeyboardInterrupt):
            wdriver.quit()
            quit()

        if isinstance(e, NoSuchElementException):
            wdriver.implicitly_wait(0)

        print_message("Can't find email address. Skipping...",
                      error=True, driver_data=data_store)

    # Find and format dsp name
    try:
        dsp_name = wdriver.find_element(
            "xpath", "//div/div/md-card[1]/div/md-card-content/div[1]/div[1]/ng-transclude/p[1]").text.replace("Business Name: ", "").strip()
    except Exception as e:
        if isinstance(e, KeyboardInterrupt):
            wdriver.quit()
            quit()

        if isinstance(e, NoSuchElementException):
            wdriver.implicitly_wait(0)

        print_message("Couldn't find DSP. Skipping...",
                      error=True, driver_data=data_store)
        dsp_name = ""

    try:
        data_store["dsp"] = DSP_MAP[dsp_name]
    except KeyError:
        if dsp_name:
            print_message("Couldn't find the DSP short code for {}. They are either a new DSP or their name changed. Please update the script to fix this error.".format(
                dsp_name), error=True, driver_data=data_store)

    # Find and format employee id
    try:
        data_store["employee_id"] = wdriver.find_element(
            "xpath", "//md-card[2]/div/md-card-content/div/div[2]/ng-transclude/p[1]").text.replace("PeopleSoft ID: ", "").replace(" [FCLM] [PhoneTool]", "").strip()
    except Exception as e:
        if isinstance(e, KeyboardInterrupt):
            wdriver.quit()
            quit()

        if isinstance(e, NoSuchElementException):
            wdriver.implicitly_wait(0)

        print_message("Couldn't find employee_id. Skipping...",
                      error=True, driver_data=data_store)

    # Onboarding progress is tricky to grab, because it loads in slower, so we just have to keep checking for it over 10 seconds
    onboarding_progress = ""
    while time.time() - before < 10:
        try:
            onboarding_progress = wdriver.find_element(
                "xpath", "//md-list/div[1]/div/div/h6[2]").text.replace(" COMPLETED", "").strip()
        except Exception as e:
            if isinstance(e, KeyboardInterrupt):
                wdriver.quit()
                quit()
            else:
                pass

        if onboarding_progress != "" and onboarding_progress != "0/0":
            break
        else:
            time.sleep(0.25)

    if time.time() - before >= 10:
        print_message("Can't find onboarding progress. Skipping...",
                      error=True, driver_data=data_store)
    else:
        if onboarding_progress == "15/17":
            data_store["onboarding_status"] = "N"
        else:
            data_store["onboarding_status"] = onboarding_progress
            print_message("DA has not finished onboarding",
                          driver_data=data_store)

    if ARGS["photos"]:
        try:
            img_element = wdriver.find_element("class name", "da-photo")

            # Use PIL library to manipulate image
            image = Image.open(io.BytesIO(img_element.screenshot_as_png))

            width = image.size[0]
            height = image.size[1]

            # Minimum size requirements for badges
            while width < 240 or height < 320:
                width *= 2
                height *= 2
                image = image.resize((width, height))

            FILENAME = "{}.png".format(data_store["da_name"])
            with open(os.path.join("photos", FILENAME), "wb") as f:
                image.save(f, format="PNG")
        except Exception as e:
            if isinstance(e, KeyboardInterrupt):
                wdriver.quit()
                quit()
            else:
                print_message(
                    "Could not save badge photo. Skipping...", error=True, driver_data=data_store)

    # Reset implicit wait
    wdriver.implicitly_wait(10)

    return data_store


def format_and_save_spreadsheet(workbook, ARGS, save_header=False, driver=None, index=-1):
    header_data = []
    driver_data = []
    sheet = workbook.active

    if save_header:
        # Add headers for every piece of data that needs to be saved, exclude data that is not being saved
        header_data = list(x for x in [
            "NAME" if ARGS["knet_link"] and ARGS["da_name"] else None,
            "KNET LINK" if ARGS["knet_link"] and not ARGS["da_name"] else None,
            "NAME" if ARGS["da_name"] and not ARGS["knet_link"] else None,
            "TRANSPORTER ID" if ARGS["transporter_id"] and ARGS["amc_link"] else None,
            "AMC LINK" if ARGS["amc_link"] and not ARGS["transporter_id"] else None,
            "TRANSPORTER ID" if ARGS["transporter_id"] and not ARGS["amc_link"] else None,
            "ID" if ARGS["employee_id"] else None,
            "USERNAME" if ARGS["username"] else None,
            "EMAIL" if ARGS["email"] else None,
            "DSP" if ARGS["dsps"] else None,
            "ONBOARDING" if ARGS["onboarding_status"] else None,
        ] if x is not None)

        # sheet.delete_rows(0)

        sheet.append(header_data)

    if driver is not None and index > -1 and not save_header:
        driver_data = list(x for x in [
            driver["da_name"] if ARGS["knet_link"] and ARGS["da_name"] else None,
            driver["knet_link"] if ARGS["knet_link"] and not ARGS["da_name"] else None,
            driver["da_name"] if ARGS["da_name"] and not ARGS["knet_link"] else None,
            driver["transporter_id"] if ARGS["transporter_id"] and ARGS["amc_link"] else None,
            driver["amc_link"] if ARGS["amc_link"] and not ARGS["transporter_id"] else None,
            driver["transporter_id"] if ARGS["transporter_id"] and not ARGS["amc_link"] else None,
            driver["employee_id"] if ARGS["employee_id"] else None,
            driver["username"] if ARGS["username"] else None,
            driver["email"] if ARGS["email"] else None,
            driver["dsp"] if ARGS["dsps"] else None,
            driver["onboarding_status"] if ARGS["onboarding_status"] else None
        ] if x is not None)

        sheet.append(driver_data)

        # Hyperlink the driver name with the knet link if both are present
        if (ARGS["knet_link"] and ARGS["da_name"]) and (driver["da_name"] and driver["knet_link"]):
            sheet.cell(row=index, column=1).value = driver["da_name"]
            sheet.cell(row=index, column=1).hyperlink = driver["knet_link"]
            sheet.cell(row=index, column=1).style = "Hyperlink"

        # Hyperlink the transporter id with the AM Console link if both are present
        if (ARGS["amc_link"] and ARGS["transporter_id"]) and (driver["transporter_id"] and driver["amc_link"]):
            header_row = list(sheet.rows)[0]

            # Find the column that contains the transporter id header
            trans_id_col = -1
            for i, cell in enumerate(header_row):
                if cell.value == "TRANSPORTER ID":
                    trans_id_col = i
                    break

            sheet.cell(row=index, column=trans_id_col +
                       1).value = driver["transporter_id"]
            sheet.cell(row=index, column=trans_id_col +
                       1).hyperlink = driver["amc_link"]
            sheet.cell(row=index, column=trans_id_col + 1).style = "Hyperlink"

    # Only save if there is an output file present
    if ARGS["output_file"]:
        try:
            workbook.save(ARGS["output_file"])
        except PermissionError as err:
            print_message(
                "Could not save data to the output file. Make sure you don't have it open while running the script.", error=True, driver_data=driver)


def print_message(message, driver_data=None, error=False):
    # Format error message
    if error:
        formatted_message = Fore.RED + Style.BRIGHT + "ERROR: " + Style.NORMAL
    else:
        formatted_message = Fore.YELLOW + Style.BRIGHT + "INFO: " + Style.NORMAL

    # Skip driver if none was provided
    if driver_data is None:
        formatted_message += "{}".format(message)
        print(formatted_message)
        return

    # Format the message with any driver info provided
    formatted_message += Fore.WHITE + "["
    if driver_data["da_name"]:
        formatted_message += Fore.BLUE + \
            driver_data["da_name"] + Fore.WHITE + "::"

    if driver_data["employee_id"]:
        formatted_message += Fore.CYAN + \
            driver_data["employee_id"] + Fore.WHITE + "::"

    if driver_data["transporter_id"]:
        formatted_message += Fore.GREEN + \
            driver_data["transporter_id"] + Fore.WHITE

    # Remove extra :: at the end
    if formatted_message[-2:] == "::":
        formatted_message = formatted_message[:-2]

    formatted_message += "] "

    if error:
        formatted_message += Fore.RED
    else:
        formatted_message += Fore.YELLOW

    formatted_message += message

    print(formatted_message)


def main():

    # Initialize colorama
    init(autoreset=True)

    ARGS = get_args_from_menu()

    # parser = argparse.ArgumentParser()
    # parser.add_argument("--in", "--input", help="The file with usernames or user ids. Only one username or id per line.", metavar="FILE", dest="input")
    # parser.add_argument("--out", "--output", help="Specify a file to save user info to. It's saved as a xlsx file.", metavar="FILE", dest="output")
    # parser.add_argument("--pass", "--password", help="The new password you want set for all users. Make sure it's unique and none of the users have used this password before.", dest="password")
    # parser.add_argument("-n", "--name", help="Save the person's name.", action="store_true")
    # parser.add_argument("-u", "--user", help="Save the person's username.", action="store_true")
    # parser.add_argument("-l", "--link", help="Save the person's profile link.", action="store_true")
    # parser.add_argument("-i", help="Save the person's ID.", action="store_true", dest="id")
    # args = parser.parse_args()

    # if args.input is None:
    #     print("Error: Please specify an input file containing user IDs.\nUse the -h option for more help.")
    #     quit()

    # if args.output is None and (args.name or args.user or args.link or args.id):
    #     print("Error: Please specify an output file for user information.\nUse the -h option for more help.")
    #     quit()

    training_roster = []
    try:
        with open(ARGS["input_file"], "r") as f:
            training_roster = f.readlines()
    except Exception:
        print_message("Could not open input file {}".format(
            ARGS["input_file"]), error=True)
        quit()

    # Throw away the first line with headers
    if (training_roster[0] == "Transporter Id,Amazon Employee Id\n"):
        training_roster = training_roster[1::]

    wdriver = webdriver.Firefox(service=Service(log_path=os.path.devnull))
    # Wait for up to 10 seconds for pages to finish loading
    wdriver.implicitly_wait(10)

    # Create photos folder
    if ARGS["photos"]:
        try:
            os.mkdir("photos")
        except FileExistsError:
            pass

    # Initialize xlsx workbook
    workbook = Workbook()

    format_and_save_spreadsheet(workbook, ARGS, save_header=True)

    for index, line in enumerate(training_roster, start=2):
        print_message("Processing driver {} of {}".format(
            index - 1, len(training_roster)))
        # Inititalize driver object
        driver = {
            "amc_link": "",
            "da_name": "",
            "dsp": "",
            "email": "",
            "employee_id": "",
            "knet_link": "",
            "onboarding_status": "",
            "transporter_id": "",
            "username": "",
        }

        # Extract and identify ids from input file
        data = line.split(",")

        for id in data:
            id = id.strip()

            if TRANSPORTER_ID_REGEX.match(id):
                driver["transporter_id"] = id
            elif EMPLOYEE_ID_REGEX.match(id):
                driver["employee_id"] = id
            elif USERNAME_REGEX.match(id):
                driver["username"] = id
            else:
                print_message(
                    "Unrecognized id: '{}'. Make sure you are only using employee and transporter ids in the input file".format(id), error=True)

        # First check AMC for info. No data will be collected if there isn't a transporter ID in the input file
        if (ARGS["da_name"] or ARGS["amc_link"] or ARGS["employee_id"] or ARGS["email"] or ARGS["dsps"] or ARGS["onboarding_status"] or ARGS["photos"]):
            amc_data = get_amc_data(wdriver, driver, ARGS)

            driver.update(amc_data)

        # If transporter ID is missing, we cannot get data on AMC, but we can still visit knet to get some data
        if (not driver["transporter_id"] and (ARGS["da_name"] or ARGS["employee_id"])) or ARGS["knet_link"] or ARGS["username"] or ARGS["reset_knet_pass"]:
            knet_data = get_knet_data_and_change_password(
                wdriver, driver, ARGS)

            driver.update(knet_data)

        format_and_save_spreadsheet(workbook, ARGS, driver=driver, index=index)

    wdriver.close()

    print("Finished!")


if __name__ == "__main__":
    main()
